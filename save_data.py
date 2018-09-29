from openpyxl import load_workbook
from openpyxl import Workbook
from utils import load_xlsx

def Data_save(name_data, type_data, source_data, Intime_data, Outtime_data, site):
	wb = load_workbook(filename = load_xlsx('simpledata.xlsx'))
	
	sheet_site = (site - 1) / 243 + 1
	
	if site % 243 == 0:
		row_site = 244
	else :
		row_site = site % 243 + 1

	ws = wb['Sheet%d'%(sheet_site)]
	ws['A%d'%(row_site)] = name_data
	ws['B%d'%(row_site)] = type_data
	ws['C%d'%(row_site)] = source_data
	ws['D%d'%(row_site)] = Intime_data
	ws['E%d'%(row_site)] = Outtime_data
	wb.save('simpledata.xlsx')
		