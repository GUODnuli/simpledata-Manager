from openpyxl import load_workbook
from openpyxl import Workbook
from utils import load_xlsx

def Openfile(site):
	wb = load_workbook(filename = load_xlsx('simpledata.xlsx'))

	sheet_site = (site - 1) / 243 + 1

	if site % 243 == 0:
		row_site = 244
	else :
		row_site = site % 243 + 1

	Ans = []
	ws = wb['Sheet%d'%(sheet_site)]

	Ans.insert(0,ws['A%d'%(row_site)].value)
	Ans.insert(1,ws['B%d'%(row_site)].value)
	Ans.insert(2,ws['C%d'%(row_site)].value)
	Ans.insert(3,ws['D%d'%(row_site)].value)
	Ans.insert(4,ws['E%d'%(row_site)].value)

	wb.save('simpledata.xlsx')
	return Ans